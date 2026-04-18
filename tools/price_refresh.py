"""
Kip It Real — weekly JustBallGloves price refresher.

This helper does all the xlsx I/O for the price refresh workflow. The
actual HTTP fetches are driven by a Claude session using WebFetch
(the sandbox cannot hit the internet directly).

Usage (called by a Claude session):

    # 1. Emit the list of rows to refresh
    python price_refresh.py extract \
        --catalog "<workspace>/Glove_CatalogV1_final_v2.xlsx" \
        --out      "<workspace>/catalog_price_refresh/<DATE>/targets.csv"

    # 2. [Claude session fills in scraped.csv by calling WebFetch per URL]

    # 3. Apply the scraped results back to a fresh copy of the xlsx and
    #    write the diff + needs-review CSVs.
    python price_refresh.py apply \
        --catalog  "<workspace>/Glove_CatalogV1_final_v2.xlsx" \
        --scraped  "<workspace>/catalog_price_refresh/<DATE>/scraped.csv" \
        --outdir   "<workspace>/catalog_price_refresh/<DATE>/"

`scraped.csv` shape (written by the Claude session, read by step 3):

    id,price,msrp,stock,error
    akadema-2026-ano315,199.99,269.99,in stock,
    wilson-wbw100795115,,,out of stock,
    rawlings-pror314-32ccb,,,,404 not found

Rules applied by `apply`:
    - If price parsed and differs from old price by > $0.50 → update price.
    - If msrp parsed → update msrp.
    - On any successful fetch → update lastVerified to --today.
    - If fetch errored OR stock in {out of stock, discontinued, backorder}
      → leave price/msrp alone, append to needs_review.csv.
    - If |new − old| > $30 → also flag as `large_delta` in price_diff.csv.

Outputs saved to --outdir:
    - Glove_CatalogV1_refreshed.xlsx
    - price_diff.csv
    - needs_review.csv
    - summary.txt
"""
from __future__ import annotations

import argparse
import csv
import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "Glove-template"
LARGE_DELTA_USD = 30.0
CHANGE_THRESHOLD_USD = 0.50
REFRESH_STOCK_STATES = {"in stock", "in-stock"}
SKIP_STOCK_STATES = {"out of stock", "out-of-stock", "backorder", "discontinued"}


def _load_sheet(path: Path):
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"[error] sheet '{SHEET_NAME}' not found in {path}")
    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    for req in ("id", "purchaseLinks", "price", "msrp", "lastVerified", "name", "brand"):
        if req not in col_idx:
            sys.exit(f"[error] required column '{req}' missing from sheet")
    return wb, ws, col_idx


def cmd_extract(args: argparse.Namespace) -> int:
    wb, ws, col = _load_sheet(Path(args.catalog))
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    with out.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "name", "brand", "old_price", "old_msrp", "url"])
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(row=r, column=col["id"]).value
            url = ws.cell(row=r, column=col["purchaseLinks"]).value
            if not rid or not url:
                continue
            url = str(url).strip()
            if not url.startswith("http"):
                continue
            w.writerow([
                rid,
                ws.cell(row=r, column=col["name"]).value or "",
                ws.cell(row=r, column=col["brand"]).value or "",
                ws.cell(row=r, column=col["price"]).value,
                ws.cell(row=r, column=col["msrp"]).value,
                url,
            ])
            rows_written += 1
    print(f"[extract] wrote {rows_written} rows → {out}")
    return 0


def _to_float(val: str) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in {"none", "null", "n/a", "not shown"}:
        return None
    s = s.replace("$", "").replace(",", "").replace(" USD", "").strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def cmd_apply(args: argparse.Namespace) -> int:
    catalog = Path(args.catalog)
    scraped = Path(args.scraped)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    # Work on a copy so the source xlsx stays untouched.
    refreshed = outdir / "Glove_CatalogV1_refreshed.xlsx"
    shutil.copyfile(catalog, refreshed)
    wb, ws, col = _load_sheet(refreshed)

    # Build id → row lookup.
    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col["id"]).value
        if rid:
            id_to_row[rid] = r

    today = args.today or date.today().isoformat()

    diffs = []          # rows where price or msrp changed
    review = []         # rows that need manual follow-up
    no_change = 0
    applied = 0

    with scraped.open() as f:
        reader = csv.DictReader(f)
        for rec in reader:
            rid = (rec.get("id") or "").strip()
            if not rid:
                continue
            row = id_to_row.get(rid)
            if row is None:
                review.append({"id": rid, "reason": "id not found in catalog"})
                continue

            err = (rec.get("error") or "").strip()
            stock = (rec.get("stock") or "").strip().lower()
            new_price = _to_float(rec.get("price"))
            new_msrp = _to_float(rec.get("msrp"))

            old_price = ws.cell(row=row, column=col["price"]).value
            old_msrp = ws.cell(row=row, column=col["msrp"]).value
            name = ws.cell(row=row, column=col["name"]).value or ""
            brand = ws.cell(row=row, column=col["brand"]).value or ""

            if err:
                review.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_price": old_price, "old_msrp": old_msrp,
                    "stock": stock, "reason": f"fetch error: {err}",
                })
                continue

            if stock in SKIP_STOCK_STATES:
                review.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_price": old_price, "old_msrp": old_msrp,
                    "stock": stock, "reason": f"stock={stock}",
                })
                continue

            if new_price is None:
                review.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_price": old_price, "old_msrp": old_msrp,
                    "stock": stock, "reason": "no price parsed",
                })
                continue

            price_changed = (
                isinstance(old_price, (int, float))
                and abs(new_price - float(old_price)) > CHANGE_THRESHOLD_USD
            ) or not isinstance(old_price, (int, float))

            msrp_changed = (
                new_msrp is not None and (
                    not isinstance(old_msrp, (int, float))
                    or abs(new_msrp - float(old_msrp)) > CHANGE_THRESHOLD_USD
                )
            )

            # Always update lastVerified on a successful fetch.
            ws.cell(row=row, column=col["lastVerified"]).value = today
            applied += 1

            if price_changed:
                ws.cell(row=row, column=col["price"]).value = new_price
            if msrp_changed:
                ws.cell(row=row, column=col["msrp"]).value = new_msrp

            if price_changed or msrp_changed:
                delta = None
                delta_pct = None
                large = False
                if isinstance(old_price, (int, float)) and old_price:
                    delta = round(new_price - float(old_price), 2)
                    delta_pct = round(delta / float(old_price) * 100, 1)
                    large = abs(delta) > LARGE_DELTA_USD
                diffs.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_price": old_price, "new_price": new_price,
                    "old_msrp": old_msrp, "new_msrp": new_msrp if msrp_changed else "",
                    "delta_usd": delta if delta is not None else "",
                    "delta_pct": delta_pct if delta_pct is not None else "",
                    "stock": stock,
                    "flag": "large_delta" if large else "",
                })
            else:
                no_change += 1

    wb.save(refreshed)

    # Write diff CSV
    diff_path = outdir / "price_diff.csv"
    with diff_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "id", "name", "brand",
            "old_price", "new_price", "old_msrp", "new_msrp",
            "delta_usd", "delta_pct", "stock", "flag",
        ])
        w.writeheader()
        for d in sorted(diffs, key=lambda x: -abs(float(x["delta_usd"] or 0))):
            w.writerow(d)

    # Write needs_review CSV
    review_path = outdir / "needs_review.csv"
    with review_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "id", "name", "brand", "old_price", "old_msrp", "stock", "reason",
        ])
        w.writeheader()
        for r in review:
            w.writerow({k: r.get(k, "") for k in w.fieldnames})

    # Summary
    summary_path = outdir / "summary.txt"
    summary = [
        f"Kip It Real — weekly price refresh ({today})",
        "─" * 52,
        f"  Rows with successful fetch:       {applied}",
        f"  Rows with price change applied:   {len(diffs)}",
        f"  Rows unchanged (within $0.50):    {no_change}",
        f"  Rows needing review (see CSV):    {len(review)}",
        f"  Large deltas (|Δ| > $30):         {sum(1 for d in diffs if d['flag'] == 'large_delta')}",
        "",
        f"  Refreshed workbook: {refreshed.name}",
        f"  Diff CSV:           {diff_path.name}",
        f"  Needs-review CSV:   {review_path.name}",
    ]
    summary_path.write_text("\n".join(summary) + "\n")
    print("\n".join(summary))
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    sub = p.add_subparsers(dest="cmd", required=True)

    e = sub.add_parser("extract", help="emit targets.csv of (id,url) rows")
    e.add_argument("--catalog", required=True)
    e.add_argument("--out", required=True)
    e.set_defaults(func=cmd_extract)

    a = sub.add_parser("apply", help="apply scraped.csv back to catalog")
    a.add_argument("--catalog", required=True)
    a.add_argument("--scraped", required=True)
    a.add_argument("--outdir", required=True)
    a.add_argument("--today", help="override today's date (YYYY-MM-DD)")
    a.set_defaults(func=cmd_apply)

    ns = p.parse_args()
    return ns.func(ns)


if __name__ == "__main__":
    raise SystemExit(main())
