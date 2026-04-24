"""
Kip It Real — JBG URL repair helper.

Broken rows in the catalog have URLs that 404 because the numeric
product-ID path segment is missing. Example:

  broken:    https://www.justballgloves.com/product/<slug>/?rfsn=<aff>
  canonical: https://www.justballgloves.com/product/<slug>/<NNNNN>/?rfsn=<aff>

This helper supports two commands used by the Claude-driven workflow:

  1. `audit`
     Scan the catalog and emit broken_urls.csv — rows where the current
     purchaseLinks URL is missing the numeric ID (or otherwise malformed).

  2. `apply`
     Read a resolved.csv produced by the Claude session (columns:
     id, canonical_url, status, reason) and update the xlsx:
        - status=fixed     → write canonical_url into purchaseLinks
        - status=unchanged → do nothing
        - status=review    → do nothing, log to needs_review
     Always preserves the ?rfsn=9019128.19c2ec affiliate param.
     Writes change_log.csv and needs_review.csv alongside the
     updated xlsx in the output directory.

Usage:

  python url_repair.py audit \
      --catalog Glove_CatalogV1_final_v2.xlsx \
      --out     catalog_url_repair/<DATE>/broken_urls.csv

  python url_repair.py apply \
      --catalog  Glove_CatalogV1_final_v2.xlsx \
      --resolved catalog_url_repair/<DATE>/resolved.csv \
      --outdir   catalog_url_repair/<DATE>/
"""
from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "Glove-template"
AFFILIATE = "?rfsn=9019128.19c2ec"
CANONICAL_RE = re.compile(
    r"^https://www\.justballgloves\.com/product/[a-z0-9\-]+/(\d+)/\?rfsn=9019128\.19c2ec$"
)
BROKEN_RE = re.compile(
    r"^https://www\.justballgloves\.com/product/([a-z0-9\-]+)/\?rfsn=9019128\.19c2ec$"
)


def _load(path: Path):
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    return wb, ws, col


def cmd_audit(args: argparse.Namespace) -> int:
    _, ws, col = _load(Path(args.catalog))
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with out.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["row", "id", "sport", "brand", "name", "sku", "slug", "current_url"])
        for r in range(2, ws.max_row + 1):
            url = (ws.cell(row=r, column=col["purchaseLinks"]).value or "").strip()
            m = BROKEN_RE.match(url)
            if not m:
                continue
            slug = m.group(1)
            sku = slug.split("--")[-1] if "--" in slug else slug
            w.writerow([
                r,
                ws.cell(row=r, column=col["id"]).value,
                ws.cell(row=r, column=col["sport"]).value,
                ws.cell(row=r, column=col["brand"]).value,
                ws.cell(row=r, column=col["name"]).value,
                sku,
                slug,
                url,
            ])
            n += 1
    print(f"[audit] {n} broken rows → {out}")
    return 0


def _normalise(url: str) -> str:
    """Ensure the affiliate param is present (and only that param)."""
    url = url.strip().rstrip(")")
    # Strip any existing query
    base = url.split("?", 1)[0]
    # Ensure trailing slash before query
    if not base.endswith("/"):
        base = base + "/"
    return base + AFFILIATE


def cmd_apply(args: argparse.Namespace) -> int:
    catalog = Path(args.catalog)
    resolved = Path(args.resolved)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    refreshed = outdir / "Glove_CatalogV1_urls_repaired.xlsx"
    shutil.copyfile(catalog, refreshed)
    wb, ws, col = _load(refreshed)

    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col["id"]).value
        if rid:
            id_to_row[rid] = r

    change_rows = []
    review_rows = []
    fixed_count = 0
    unchanged_count = 0
    review_count = 0

    with resolved.open() as f:
        reader = csv.DictReader(f)
        for rec in reader:
            rid = (rec.get("id") or "").strip()
            status = (rec.get("status") or "").strip().lower()
            new_url = (rec.get("canonical_url") or "").strip()
            reason = (rec.get("reason") or "").strip()
            if not rid:
                continue
            row = id_to_row.get(rid)
            if row is None:
                review_rows.append({
                    "id": rid,
                    "name": rec.get("name", ""),
                    "brand": rec.get("brand", ""),
                    "old_url": rec.get("old_url", ""),
                    "reason": "id not found in catalog",
                })
                review_count += 1
                continue
            old_url = (ws.cell(row=row, column=col["purchaseLinks"]).value or "").strip()
            name  = ws.cell(row=row, column=col["name"]).value or ""
            brand = ws.cell(row=row, column=col["brand"]).value or ""

            if status == "fixed":
                norm = _normalise(new_url)
                if not CANONICAL_RE.match(norm):
                    # Safety — resolver returned something odd; flag for review
                    review_rows.append({
                        "id": rid, "name": name, "brand": brand,
                        "old_url": old_url,
                        "reason": f"proposed url failed canonical check: {new_url}",
                    })
                    review_count += 1
                    continue
                if norm == old_url:
                    unchanged_count += 1
                    change_rows.append({
                        "id": rid, "name": name, "brand": brand,
                        "old_url": old_url, "new_url": norm,
                        "status": "unchanged", "reason": "already canonical",
                    })
                    continue
                ws.cell(row=row, column=col["purchaseLinks"]).value = norm
                fixed_count += 1
                change_rows.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_url": old_url, "new_url": norm,
                    "status": "fixed", "reason": reason or "canonical url resolved",
                })
            elif status == "review":
                review_rows.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_url": old_url,
                    "reason": reason or "manual review requested",
                })
                review_count += 1
            else:  # unchanged
                unchanged_count += 1
                change_rows.append({
                    "id": rid, "name": name, "brand": brand,
                    "old_url": old_url, "new_url": old_url,
                    "status": "unchanged", "reason": reason or "no change",
                })

    wb.save(refreshed)

    # Write change log
    change_log = outdir / "change_log.csv"
    with change_log.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "id", "name", "brand", "old_url", "new_url", "status", "reason",
        ])
        w.writeheader()
        w.writerows(change_rows)

    # Write review list
    review_path = outdir / "needs_review.csv"
    with review_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "id", "name", "brand", "old_url", "reason",
        ])
        w.writeheader()
        w.writerows(review_rows)

    summary = [
        "URL repair summary",
        "─" * 40,
        f"  Fixed:    {fixed_count}",
        f"  Unchanged:{unchanged_count}",
        f"  Review:   {review_count}",
        "",
        f"  Workbook:    {refreshed.name}",
        f"  Change log:  {change_log.name}",
        f"  Review list: {review_path.name}",
    ]
    (outdir / "summary.txt").write_text("\n".join(summary) + "\n")
    print("\n".join(summary))
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    sub = p.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("audit")
    a.add_argument("--catalog", required=True)
    a.add_argument("--out", required=True)
    a.set_defaults(func=cmd_audit)

    ap = sub.add_parser("apply")
    ap.add_argument("--catalog", required=True)
    ap.add_argument("--resolved", required=True)
    ap.add_argument("--outdir", required=True)
    ap.set_defaults(func=cmd_apply)

    ns = p.parse_args()
    return ns.func(ns)


if __name__ == "__main__":
    raise SystemExit(main())
